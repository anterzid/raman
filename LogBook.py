__author__ = "Anastasia Terzidou"
__copyright__ = "Copyright 2019, The EasyRaman Project"
__credits__ = ["Anastasia Terzidou"]
__version__ = "1.0 beta"
__maintainer__ = "Anastasia Terzidou"
__email__ = "anterzid@gmail.com"
__status__ = "Development"



from tkinter import *
from openpyxl import load_workbook
from openpyxl import Workbook
import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import *
from tkinter import messagebox
import os
import tkinter.ttk

cwp=os.path.abspath(os.path.dirname(__file__))
filename="Raman Logbook.xlsx"

try:																#detects if the excel file already exists and if it doesn't it creates it
	wb=Workbook(filename)
	ws=wb.active
except:
	wb=Workbook()
	ws=wb.active
	first_list=["Measurement Code","User", "Laser Line","Laser Power","Filter","Sample","Kind","Spectrometer","Lens","Hole","Grating","Acquisition Time","Repetitions","Refererence Neon",
				"Actual Neon","Comments","Ruby in before", "Ruby air before", "Pressure Before", "Ruby in after", "Ruby out after", "Pressure After","Mean Pressure"]

	for col, val in enumerate(first_list, start=1):
		ws.cell(row=1, column=col).value = val

	wb.save(filename)


def saveData():														#a new function which appends the data given by the user in an Excel file by the help of Pandas DataFrames

	try:
		units=unit.get()
		lasers=laser.get()
		datecode=datetime.today().strftime('%y%m%d')
		mcodes= datecode + mcode.get()
		susers=suser.get()
		snames=sname.get()
		skinds=skind.get()
		scenters=scenter.get()
		samplepowr=samplepow.get()
		sholes=shole.get()
		sfilters=sfilter.get()
		sgratings=sgrating.get()
		slenses=slens.get()
		acqutimes=acqutime.get()
		repses=reps.get()
		comments=comment.get("1.0",'end-1c')
		rubyinbefore=rubinb.get()
		rubyoutbefore=ruboutb.get()
		rubyinafter=rubinaf.get()
		rubyoutafter=ruboutaf.get()
		refneons=refneon.get()
		corrneons=corrneon.get()
		date=datetime.today().strftime('%m/%y')

		wb=load_workbook(filename)
		ws = wb.active

		datalist=[mcodes,susers,lasers,samplepowr,sfilters,snames,skinds,scenters,slenses,sholes,sgratings,acqutimes,repses,refneons,corrneons
					,comments,rubyinbefore,rubyoutbefore,pressb,rubyinafter,rubyoutafter,pressa,pressm,units]



		df=pd.DataFrame(data=datalist)
		dfTrans=df.T																#transpose the above list


		for r in dataframe_to_rows(dfTrans, index=False, header=False):
			ws.append(r)


		button_pressed()

		wb.save(filename)


	except:
		messagebox.showwarning("Warning","Please close the Excel File and press the 'Save Data' button again!")


def button_pressed():
	# put text
	saved=Label(master, text = "Saved!",bg="#8fbc8f",font='Lucida-Handwriting 11 bold').grid (row=82, column=5)
	# run clear_label after 2000ms (2s)
	master.after(2000, clear_label)

def clear_label():
	# remove text
	saved=Label(master, text = "                   ",bg="#8fbc8f").grid (row=61, column=5)

def date_button():
	Label(master, text = date.today(),bg="#8fbc8f",font="Calibri 12 bold").grid (row=1, column=3)



#initialize GUI
master = Tk()

master.title("Raman Logbook v.1.5 beta")  								  			 											#GUI's name
master.configure(background="#8fbc8f")
master.geometry('1500x1000')

#General info grid

Label(master, text="GENERAL",bg="#8fbc8f", font=('Calibri 16 bold underline')).grid(row=0, column=4)  							#title of grid

unit = StringVar(master)
unit.set("LabRam HR")                                                       													#set the initial parameter in the option menu
Label(master, text="Unit",bg="#8fbc8f",font=('Lucida-Handwriting 11')).grid(row=1)												#label next to the menu and at the same time placing it in the grid
unitmenu = OptionMenu(master, unit, "LabRam HR", "T64000")
unitmenu.config(bg="#8fbc8f")
unitmenu.grid(row=1, column=1)

Label(master, text = "Date",bg="#8fbc8f",font=('Lucida-Handwriting 11')).grid (row=1, column=2)
Label(master, text = date.today(),bg="#8fbc8f",font="Calibri 12 bold").grid (row=1, column=3)
Button(master,text="Re",bg="#8fbc8f",font='Lucida-Handwriting 11 bold',command=date_button).grid(row=1,column=4)


laser=StringVar(master)
laser.set("Cobolt 514.63 nm")
Label(master, text = "Laser Line",bg="#8fbc8f",font=('Lucida-Handwriting 11')).grid (row=1,column=5,pady=25)
lasermenu = OptionMenu(master, laser, "Cobolt 514.63 nm", "Red 632.8 nm", "Cobolt 08-NLD 784.8308 nm")
lasermenu.config(bg="#8fbc8f")
lasermenu.grid(row=1, column=6)


#Sample Data Grid

Label(master, text="SAMPLE DATA",bg="#8fbc8f", font=('Calibri 16 bold underline')).grid(row=2,column=4)


Label(master, text = datetime.today().strftime('%y%m%d'),bg="#8fbc8f",font=('Calibri 12 bold')).grid (row=3)
mcode=Entry(master)
mcode.config(bg="#d3d3d3")
mcode.grid(row=3,column=1)

Label(master, text = "User",bg="#8fbc8f",font=('Lucida-Handwriting 11')).grid (row=3, column=2)
suser=Entry(master)
suser.config(bg="#d3d3d3")
suser.grid(row=3, column=3)

Label(master, text = "Sample's Name",bg="#8fbc8f",font=('Lucida-Handwriting 11')).grid (row=3,column=4)
sname=Entry(master)
sname.config(bg="#d3d3d3")
sname.grid(row=3,column=5)

skind=StringVar(master)
Label(master, text = "Sample's Kind",bg="#8fbc8f",font=('Lucida-Handwriting 11')).grid (row=3,column=6)
skindmenu=OptionMenu(master,skind,"solid","powder","liquid")
skindmenu.config(bg="#8fbc8f")
skindmenu.grid(row=3,column=7)


Label(master, text="",bg="#8fbc8f").grid(row=4, rowspan=2)


Label(master, text = "Spectrometer Center",bg="#8fbc8f",font=('Lucida-Handwriting 11')).grid (row=7)
scenter=Entry(master)
scenter.config(bg="#d3d3d3")
scenter.grid(row=7,column=1)

Label(master, text = "Laser Power on sample",bg="#8fbc8f",font=('Lucida-Handwriting 11')).grid(row=7,column=2)
samplepow=Entry(master)
samplepow.config(bg="#d3d3d3")
samplepow.grid(row=7,column=3)


Label(master, text="",bg="#8fbc8f").grid(row=8, rowspan=3)


Label(master, text = "Hole",bg="#8fbc8f",font=('Lucida-Handwriting 11')).grid (row=12)
shole=Entry(master)
shole.config(bg="#d3d3d3")
shole.grid(row=12,column=1)

sfilter=StringVar(master)
Label(master, text = "Filter",bg="#8fbc8f",font=('Lucida-Handwriting 11')).grid (row=12, column=2)
sfiltermenu = OptionMenu(master, sfilter, "D 0.3", "D 0.6", "D 1" ,"D 2","D 3","D 4")
sfiltermenu.config(bg="#8fbc8f")
sfiltermenu.grid(row=12,column=3)


Label(master, text="",bg="#8fbc8f").grid(row=13, rowspan=3)



sgrating=StringVar(master)
Label(master, text = "Grating",bg="#8fbc8f",font=('Lucida-Handwriting 11')).grid (row=17)
sgratingmenu = OptionMenu(master, sgrating, "300", "600", "1800")
sgratingmenu.config(bg="#8fbc8f")
sgratingmenu.grid(row=17,column=1)

slens=StringVar(master)
Label(master, text = "Lens",bg="#8fbc8f",font=('Lucida-Handwriting 11')).grid (row=17, column=2)
slensmenu = OptionMenu(master, slens, "10x", "20x", "50x","50x SLWD","100x","100x SLWD")
slensmenu.config(bg="#8fbc8f")
slensmenu.grid(row=17,column=3)

Label(master, text = "Comments",bg="#8fbc8f",font=('Lucida-Handwriting 11')).grid (row=17,column=4)
comment=Text(master,height=7, width=35)
comment.config(bg="#d3d3d3")
comment.grid(row=17,column=5,rowspan=3,columnspan=2)


Label(master, text="",bg="#8fbc8f").grid(row=18, rowspan=3)



Label(master, text = "Acquisition Time",bg="#8fbc8f",font=('Lucida-Handwriting 11')).grid (row=22)
acqutime=Entry(master)
acqutime.config(bg="#d3d3d3")
acqutime.grid(row=22,column=1)

Label(master, text = "Repetitions",bg="#8fbc8f",font=('Lucida-Handwriting 11')).grid (row=22, column=2)
reps=Entry(master)
reps.config(bg="#d3d3d3")
reps.grid(row=22,column=3)


Label(master, text="",bg="#8fbc8f").grid(row=23, rowspan=3)


Label(master, text = "Reference Neon Line",bg="#8fbc8f",font=('Lucida-Handwriting 11')).grid (row=27)
refneon=Entry(master)
refneon.config(bg="#d3d3d3")
refneon.grid(row=27,column=1)

Label(master, text = "Corrected Neon Line",bg="#8fbc8f",font=('Lucida-Handwriting 11')).grid (row=27, column=2)
corrneon=Entry(master)
corrneon.config(bg="#d3d3d3")
corrneon.grid(row=27,column=3)


#Pressure Experiment Grid

def on_entry_click(event):
	if rubinb.cget('fg')=='#778899':
		rubinb.delete(0,"end")
		rubinb.insert(0,"")
		rubinb.config(fg="black")


def on_focusout(event):
	if rubinb.get()=="":
		rubinb.insert(0,"Numbers in 0.0 format")
		rubinb.config(fg="#778899")



pressa=0
pressb=0
pressm=0

Label(master, text="",bg="#8fbc8f").grid(row=28, rowspan=2)

Label(master, text="PRESSURE PARAMETERS",bg="#8fbc8f", font=('Calibri 16 bold underline')).grid(row=31,column=4)
tkinter.ttk.Separator(master).place(x=1, y=638, relwidth=1)
Label(master, text="",bg="#8fbc8f").grid(row=32, rowspan=2)


Label(master, text = "Ruby in",bg="#8fbc8f",font=('Lucida-Handwriting 11')).grid (row=35)
rubinb=Entry(master)
rubinb.config(bg="#d3d3d3",fg="#778899")
rubinb.insert(0,"Numbers in 0.0 format")
rubinb.bind('<FocusIn>',on_entry_click)
rubinb.bind('<FocusOut>',on_focusout)
rubinb.grid(row=35,column=1)


Label(master, text = "Ruby air",bg="#8fbc8f",font=('Lucida-Handwriting 11')).grid (row=35,column=2)
ruboutb=Entry(master)
ruboutb.config(bg="#d3d3d3")
ruboutb.grid(row=35,column=3)


def pressurebefore():

	global pressb

	pressb=float(1.33*(float(rubinb.get())-float(ruboutb.get())))
	Label(master, text = pressb,bg="#8fbc8f",font=('Lucida-Handwriting 11')).grid (row=35,column=5)


buttonpressb= Button(master, text="Calculate P before (kbar)",bg="#8fbc8f",font=('Lucida-Handwriting 11 bold'), command=pressurebefore)
buttonpressb.grid(row=35,column=4)


Label(master, text="",bg="#8fbc8f").grid(row=36, rowspan=2)


Label(master, text = "Ruby in",bg="#8fbc8f",font=('Lucida-Handwriting 11')).grid (row=39)
rubinaf=Entry(master)
rubinaf.config(bg="#d3d3d3")
rubinaf.grid(row=39,column=1)


Label(master, text = "Ruby air",bg="#8fbc8f",font=('Lucida-Handwriting 11')).grid (row=39,column=2)
ruboutaf=Entry(master)
ruboutaf.config(bg="#d3d3d3")
ruboutaf.grid(row=39,column=3)


def pressureafter():

	global pressa
	global pressm

	pressa=float(1.33*(float(rubinaf.get())-float(ruboutaf.get())))
	Label(master, text = pressa,bg="#8fbc8f",font=('Lucida-Handwriting 11')).grid (row=39,column=5)
	pressm=(pressb+pressa)/2
	Label(master, text = pressm,bg="#8fbc8f",font=('Lucida-Handwriting 11 bold')).grid (row=43,column=5)


buttonpressa= Button(master, text="Calculate P after (kbar)",bg="#8fbc8f",font=('Lucida-Handwriting 11 bold'), command=pressureafter)
buttonpressa.grid(row=39,column=4)




Label(master, text="",bg="#8fbc8f").grid(row=40, rowspan=2)

Label(master, text = "P mean",bg="#8fbc8f",font=('Lucida-Handwriting 11 bold')).grid (row=43,column=4)

Label(master, text="",bg="#8fbc8f").grid(row=45, rowspan=4)


Label(master, text="",bg="#8fbc8f").grid(row=50, rowspan=10,sticky="s")
buttondata= Button(master, text="Save Data",bg="#8fbc8f",font=('Lucida-Handwriting 11 bold'), command=saveData)					#button which calls the function to save the data to the Excel file
buttondata.grid(row=61,column=4)


mainloop()
